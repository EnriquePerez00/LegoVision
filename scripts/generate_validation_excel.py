# -*- coding: utf-8 -*-
# scripts/generate_validation_excel.py
import os, sys, json, argparse, urllib.request
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(project_root)
FACE_NAMES = {0: 'Top', 1: 'Side', 2: 'Bottom'}

PART_TYPE_MAP = {
    '3001': 'Brick 2x4',
    '3002': 'Brick 2x3',
    '3003': 'Brick 2x2',
    '3004': 'Brick 1x2',
    '3005': 'Brick 1x1',
    '3010': 'Brick 1x4',
    '3020': 'Plate 2x4',
    '3021': 'Plate 2x3',
    '3022': 'Plate 2x2',
    '3023': 'Plate 1x2',
    '3024': 'Plate 1x1',
    '3037': 'Slope 45 2x4',
    '3039': 'Slope 45 2x2',
    '3062': 'Brick Round 1x1',
    '3068': 'Tile 2x2',
    '3069': 'Tile 1x2',
    '3298': 'Slope 33 3x2',
    '3622': 'Brick 1x3',
    '3665': 'Slope Inverted 45 2x1',
    '3700': 'Technic Brick 1x2',
    '3701': 'Technic Brick 1x4',
    '3710': 'Plate 1x4',
    '2412': 'Tile Grille 1x2',
    '2420': 'Plate Corner 2x2',
    '2431': 'Tile 1x4',
    '2432': 'Tile 1x2 with Handle',
    '2877': 'Brick Profile 1x2',
    '4032': 'Plate Round 2x2',
    '4070': 'Brick 1x1 Headlight',
    '6141': 'Plate Round 1x1',
    '6636': 'Tile 1x6',
    '11477': 'Slope Curved 2x1',
    '15068': 'Slope Curved 2x2',
    '15573': 'Plate Modified 1x2',
    '32000': 'Technic Brick 1x2 Holes',
    '48336': 'Plate Modified 1x2 Handle',
    '54200': 'Slope 30 1x1',
    '59900': 'Cone 1x1',
    '60478': 'Plate Modified 1x2 Handle',
    '85984': 'Slope 31 1x2',
    '98138': 'Tile Round 1x1',
    '99206': 'Plate Modified 2x2',
}


def get_piece_type(ref, name_from_json=""):
    if ref in PART_TYPE_MAP: return PART_TYPE_MAP[ref]
    if name_from_json and name_from_json not in ("Pieza Lego", ""):
        return name_from_json
    try:
        sys.path.insert(0, os.path.join(project_root, "scratch"))
        from generate_synthetic_set import get_ldraw_part_path
        path = get_ldraw_part_path(ref)
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8", errors="ignore") as fh:
                for ln in fh:
                    ln = ln.strip()
                    if ln.startswith("0 ") and len(ln) > 2:
                        desc = ln[2:].strip()
                        if desc and not desc.startswith("FILE") and not desc.startswith("NOFILE"):
                            return desc
    except Exception: pass
    return "Part " + ref


def download_bricklink_image(ref, color_code, timeout=8):
    import ssl
    # Bypass SSL verification for macOS Python cert store issues
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    urls = [
        "https://img.bricklink.com/ItemImage/PN/" + str(color_code) + "/" + ref + ".png",
        "https://img.bricklink.com/ItemImage/PL/" + ref + ".png",
    ]
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
        "Referer": "https://www.bricklink.com/",
    }
    for url in urls:
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                data = resp.read()
                if len(data) > 500: return data
        except Exception: continue
    return None


_semantic_cache = {}

def load_semantic_cache(semantic_json_path):
    if not _semantic_cache and semantic_json_path and os.path.exists(semantic_json_path):
        with open(semantic_json_path) as _f:
            data = json.load(_f)
        for r in data.get("results", []):
            _semantic_cache[r["part_ref"]] = r

def get_semantic_poses_from_db(ref, semantic_json_path=None):
    # Use semantic algorithm JSON if available
    load_semantic_cache(semantic_json_path)
    if ref in _semantic_cache:
        sem = _semantic_cache[ref]
        poses = sem.get("stable_poses", [])
        return [(p["face_class"], p["area"]) for p in poses], sem.get("n_poses", len(poses))
    # Fallback: try BD
    try:
        from core.db import supabase_client
        conn = supabase_client.get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT stable_face FROM piece_embeddings WHERE part_ref = %s ORDER BY stable_face", (ref,))
            rows = cur.fetchall()
        conn.close()
        face_map = {0: "Top", 1: "Side", 2: "Bottom"}
        poses = [(face_map.get(r["stable_face"], "Side"), 0) for r in rows]
        return poses, len(poses)
    except Exception as e:
        print("  [WARN] BD/Algoritmo no disponible para " + ref + ": " + str(e))
        return [], 0


def embed_img(ws, img_path, col, row, size=80):
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        xl = XLImage(img_path)
        xl.width = size
        xl.height = size
        xl.anchor = get_column_letter(col) + str(row)
        ws.add_image(xl)
        return True
    except Exception as e:
        print("  [WARN] Error embebiendo imagen: " + str(e))
        return False


def generate_excel(set_id, validation_json_path, output_excel_path, renders_dir, semantic_json_path=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl no instalado: pip install openpyxl pillow")
        return False
    if not os.path.exists(validation_json_path):
        print("[ERROR] JSON no encontrado: " + validation_json_path)
        return False
    with open(validation_json_path, "r", encoding="utf-8") as fh:
        val_data = json.load(fh)
    report = val_data.get("report", [])
    print("[Excel] Generando para " + str(len(report)) + " piezas del set " + set_id + "...")
    sys.path.insert(0, os.path.join(project_root, "database"))
    from set_catalog import REAL_SETS
    set_info = REAL_SETS.get(set_id, {})
    cc_map = {p["ref"]: str(p.get("color_code", "0")) for p in set_info.get("parts", [])}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validacion Posiciones Estables"
    hf = Font(bold=True, color="FFFFFF", size=10)
    fs = PatternFill("solid", fgColor="1F4E79")
    fe = PatternFill("solid", fgColor="375623")
    fm = PatternFill("solid", fgColor="2E2E2E")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    disc_fill = PatternFill("solid", fgColor="FFC7CE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    COLS = [
        ("Referencia", fm, 12),
        ("Imagen BrickLink", fm, 14),
        ("Tipo de Pieza", fm, 24),
        ("# Poses Semantico (BD)", fs, 11),
        ("Orientaciones Semantico", fs, 22),
        ("# Poses Experimental", fe, 11),
        ("Orientaciones Experimental", fe, 22),
        ("Renders Poses Simuladas", fe, 55),
        ("Discrepancia", fm, 14),
    ]
    ncols = len(COLS)
    ws.merge_cells("A1:" + get_column_letter(ncols) + "1")
    tc = ws["A1"]
    tc.value = "Validacion Posiciones Estables - Set " + set_id + " - " + set_info.get("name", "")
    tc.font = Font(bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1A1A2E")
    tc.alignment = ca
    ws.row_dimensions[1].height = 22
    for ci, (ht, fi, cw) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci)
        c.value = ht; c.font = hf; c.fill = fi; c.alignment = ca; c.border = tb
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[2].height = 30
    tmp_dir = os.path.join(project_root, "data", "tmp", "excel_imgs")
    os.makedirs(tmp_dir, exist_ok=True)
    cur_row = 3
    for i, item in enumerate(report):
        ref = item["part_ref"]
        exp_poses = item.get("poses", [])
        exp_faces = item.get("experimental_faces", [])
        disc = item.get("discrepancy", False)
        sem_poses, sem_n = get_semantic_poses_from_db(ref, semantic_json_path)
        ptype = get_piece_type(ref, item.get("name", ""))
        cc = cc_map.get(ref, "0")
        ws.row_dimensions[cur_row].height = 72
        rf = alt_fill if i % 2 == 0 else None
        def sc(col, val, fo=None, _r=cur_row, _rf=rf):
            c2 = ws.cell(row=_r, column=col)
            c2.value = val; c2.alignment = ca; c2.border = tb
            if fo: c2.fill = fo
            elif _rf: c2.fill = _rf
        sc(1, ref)
        bl_path = os.path.join(tmp_dir, "bl_" + ref + "_" + cc + ".png")
        if not os.path.exists(bl_path):
            print("  Descargando BrickLink imagen para " + ref + "...")
            d = download_bricklink_image(ref, cc)
            if d:
                with open(bl_path, "wb") as fh2: fh2.write(d)
            else: bl_path = None
        if bl_path and os.path.exists(bl_path):
            embed_img(ws, bl_path, 2, cur_row, 80)
            ws.cell(row=cur_row, column=2).border = tb
            if rf: ws.cell(row=cur_row, column=2).fill = rf
        else: sc(2, "(sin imagen)")
        sc(3, ptype)
        sc(4, sem_n if sem_n > 0 else "N/A")
        sl = " / ".join(str(p[0]) for p in sem_poses) if sem_poses else "No indexado"
        sc(5, sl)
        sc(6, len(set(exp_faces)))
        el = " / ".join(FACE_NAMES.get(f, str(f)) for f in sorted(set(exp_faces)))
        sc(7, el)
        rimgs = []
        for pidx, pose in enumerate(exp_poses):
            fid = pose["face"]
            fn = "validation_" + ref + "_pose" + str(pidx) + "_face" + str(fid) + ".png"
            fp = os.path.join(renders_dir, fn)
            if os.path.exists(fp): rimgs.append(fp)
        if rimgs:
            ws.column_dimensions[get_column_letter(8)].width = max(55, len(rimgs) * 13)
            for ri, rp in enumerate(rimgs):
                ac = 8 + ri
                ws.column_dimensions[get_column_letter(ac)].width = 13
                embed_img(ws, rp, ac, cur_row, 70)
                ws.cell(row=cur_row, column=ac).border = tb
                if rf: ws.cell(row=cur_row, column=ac).fill = rf
        else:
            sc(8, "(sin renders - ejecutar Blender primero)")
        disc_col = 9 + max(0, len(rimgs) - 1) if rimgs else 9
        c_disc = ws.cell(row=cur_row, column=disc_col)
        c_disc.value = "DISCREPANCIA" if disc else "OK"
        c_disc.fill = disc_fill if disc else ok_fill
        c_disc.alignment = ca; c_disc.border = tb
        c_disc.font = Font(bold=True, color="9C0006" if disc else "375623")
        cur_row += 1
        print("  [" + str(i+1) + "/" + str(len(report)) + "] " + ref + " -> " + ("DISC" if disc else "OK"))
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Excel] Guardado en: " + output_excel_path)
    return True


def main():
    parser = argparse.ArgumentParser(description="Genera Excel comparativo de posiciones estables")
    parser.add_argument("--set_id", type=str, default="75078-1")
    parser.add_argument("--validation_json", type=str,
                        default=os.path.join(project_root, "data", "tmp", "stability_validation_results.json"))
    parser.add_argument("--output", type=str,
                        default=os.path.join(project_root, "data", "validation_report_75078.xlsx"))
    parser.add_argument("--renders_dir", type=str,
                        default=os.path.join(project_root, "data", "validation_renders"))
    parser.add_argument("--semantic_json", type=str,
                        default=os.path.join(project_root, "data", "tmp", "semantic_poses_750781.json"))
    args = parser.parse_args()
    ok = generate_excel(args.set_id, args.validation_json, args.output, args.renders_dir, args.semantic_json)
    if ok:
        print("[Excel] Completado. Abre: " + args.output)
    else:
        print("[Excel] Fallo al generar el Excel.")


if __name__ == "__main__":
    main()
