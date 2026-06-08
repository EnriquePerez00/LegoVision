# -*- coding: utf-8 -*-
"""
generate_poses_excel.py
=======================
Genera un fichero Excel en `2camaras_pieza_unica/reports/` con una fila por
cada pieza (REGULAR + MINIFIG) del set 75078-1.

Datos: leídos directamente de la BD local (PostgreSQL):
    - Inventario regular           → tabla `lego_set_parts`
    - Piezas que componen minifigs → tabla `minifig_parts`
                                     (asociadas a las minifigs en
                                     `lego_set_minifigures` para el set)
    - Conteo de poses estables     → tabla `stable_poses`

Columnas del Excel:
    1. Imagen           (PNG incrustado, descargado de BrickLink/Rebrickable)
    2. Código           (part_ref tal como aparece en BrickLink)
    3. Tipo             ("Regular" / "Minifig (sw0578)")
    4. Color            ("BL_id — Color name")
    5. Nombre           (descripción de la pieza desde el catálogo BrickLink)
    6. Poses estables   (count de stable_poses con is_stable=TRUE para esa ref
                         y set_id 75078-1)
    7. Esperadas        (vacía, para rellenar manualmente)

Uso:
    .venv/bin/python 2camaras_pieza_unica/scripts/generate_poses_excel.py
"""

import os
import sys
import io
import json

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)            # 2camaras_pieza_unica/
LEGO_ROOT    = os.path.dirname(PROJECT_ROOT)          # repo raíz

sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, PROJECT_ROOT)
sys.path.insert(0, LEGO_ROOT)
sys.path.insert(0, os.path.join(PROJECT_ROOT, "database"))

import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from supabase_client import get_connection   # type: ignore


SET_ID    = "75078-1"
OUT_DIR   = os.path.join(PROJECT_ROOT, "reports")
OUT_FILE  = os.path.join(OUT_DIR, f"poses_estables_{SET_ID}.xlsx")
CACHE_DIR = os.path.join(OUT_DIR, "_bricklink_cache")
IMG_HEIGHT = 80   # px

COLOR_CATALOG_PATH = os.path.join(PROJECT_ROOT, "database", "color_catalog.json")


# ── Mapas BrickLink y LDraw a partir de color_catalog.json ─────────────────────
NAME_TO_BL_ID:    dict = {}
NAME_TO_LDRAW_ID: dict = {}
BL_TO_LDRAW:      dict = {}
try:
    with open(COLOR_CATALOG_PATH, "r", encoding="utf-8") as fh:
        _cat = json.load(fh)
        for bl_id, info in _cat.items():
            nm  = (info.get("name") or "").strip().lower()
            ldr = str(info.get("ldraw_code") or "").strip()
            if nm and nm not in NAME_TO_BL_ID:
                NAME_TO_BL_ID[nm] = bl_id
            if nm and ldr and nm not in NAME_TO_LDRAW_ID:
                NAME_TO_LDRAW_ID[nm] = ldr
            if ldr:
                BL_TO_LDRAW[bl_id] = ldr
except Exception as e:
    print(f"[WARN] No se pudo leer color_catalog.json: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  BD: lecturas
# ══════════════════════════════════════════════════════════════════════════════

def fetch_set_inventory(set_id: str) -> dict:
    """
    Devuelve:
        {
            "set_name": str,
            "regular_parts": [ {ref, color_code, color_hex, color_name,
                                qty, name, source: 'regular',
                                minifig_ref: str|None}, ... ],
            "minifig_parts": [ {ref, color_code, color_hex, color_name,
                                qty, name, source: 'minifig',
                                minifig_ref: str}, ... ],
        }
    """
    out = {"set_name": "", "regular_parts": [], "minifig_parts": []}
    with get_connection() as conn, conn.cursor() as cur:
        cur.execute("SELECT name FROM lego_sets WHERE code = %s", (set_id,))
        row = cur.fetchone()
        out["set_name"] = row["name"] if row else f"Set {set_id}"

        # Inventario regular del set (puede incluir piezas marcadas como
        # is_minifig_part=TRUE si aparecen también en una minifig — no es
        # nuestro caso para 75078-1 pero el código lo soporta).
        cur.execute("""
            SELECT part_ref, color_code, color_hex, color_name, qty,
                   COALESCE(is_minifig_part, FALSE) AS is_minifig_part,
                   minifig_ref
            FROM lego_set_parts
            WHERE set_code = %s
            ORDER BY part_ref, color_code
        """, (set_id,))
        for r in cur.fetchall():
            out["regular_parts"].append({
                "ref":         r["part_ref"],
                "color_code":  r["color_code"],
                "color_hex":   r["color_hex"],
                "color_name":  r["color_name"],
                "qty":         r["qty"],
                "name":        "",   # se rellena más abajo si conocemos
                "source":      "regular",
                "minifig_ref": r["minifig_ref"] if r["is_minifig_part"] else None,
            })

        # Minifiguras del set
        cur.execute("""
            SELECT minifig_ref, name, qty
            FROM lego_set_minifigures
            WHERE set_code = %s
            ORDER BY minifig_ref
        """, (set_id,))
        minifigs = [dict(r) for r in cur.fetchall()]

        # Piezas de cada minifig (de la tabla minifig_parts)
        for mf in minifigs:
            mref = mf["minifig_ref"]
            cur.execute("""
                SELECT part_ref, color_code, color_hex, color_name, qty, name
                FROM minifig_parts
                WHERE minifig_ref = %s
                ORDER BY part_ref, color_code
            """, (mref,))
            for r in cur.fetchall():
                out["minifig_parts"].append({
                    "ref":         r["part_ref"],
                    "color_code":  r["color_code"],
                    "color_hex":   r["color_hex"],
                    "color_name":  r["color_name"],
                    "qty":         r["qty"],
                    "name":        r["name"] or "",
                    "source":      "minifig",
                    "minifig_ref": mref,
                })

    return out


def fetch_pose_counts(set_id: str) -> dict:
    """Devuelve {part_ref: nº de poses estables} para is_stable=TRUE."""
    counts = {}
    try:
        with get_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT part_ref, COUNT(*) AS n
                FROM stable_poses
                WHERE set_id = %s AND is_stable = TRUE
                GROUP BY part_ref
                """,
                (set_id,),
            )
            for r in cur.fetchall():
                counts[r["part_ref"]] = int(r["n"])
    except Exception as e:
        print(f"[WARN] No se pudo consultar stable_poses: {e}")
    return counts


# ══════════════════════════════════════════════════════════════════════════════
#  BrickLink / Rebrickable: descarga de imágenes
# ══════════════════════════════════════════════════════════════════════════════

BRICKLINK_URL          = "https://img.bricklink.com/ItemImage/PN/{color}/{ref}.png"
REBRICKABLE_LDRAW_URL  = "https://cdn.rebrickable.com/media/parts/ldraw/{color}/{ref}.png"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    )
}

# Códigos populares para fallback (BrickLink IDs reales)
FALLBACK_BL_IDS = ["11", "85", "86", "5", "1", "9", "8", "10", "2", "3"]


def _try_download(url: str, dst: str) -> bool:
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 200 and r.content and len(r.content) > 200:
            with open(dst, "wb") as f:
                f.write(r.content)
            return True
    except Exception as e:
        print(f"   [WARN] Descarga falló ({url}): {e}")
    return False


def download_bricklink_image(ref: str, color_code: str, color_name: str = "") -> str:
    """
    Descarga (con cache) la imagen de la pieza en BrickLink (con fallback
    a Rebrickable LDraw render). Retorna ruta local o None.
    """
    os.makedirs(CACHE_DIR, exist_ok=True)
    fname = f"{ref}_{color_code}.png"
    path  = os.path.join(CACHE_DIR, fname)
    if os.path.exists(path) and os.path.getsize(path) > 0:
        return path

    # 1) BrickLink con candidatos de color
    bl_candidates: list = []
    if color_code:
        bl_candidates.append(str(color_code))
    nm_low = (color_name or "").strip().lower()
    if nm_low and nm_low in NAME_TO_BL_ID:
        bl_id = NAME_TO_BL_ID[nm_low]
        if bl_id not in bl_candidates:
            bl_candidates.append(bl_id)
    for fb in FALLBACK_BL_IDS:
        if fb not in bl_candidates:
            bl_candidates.append(fb)

    for cand in bl_candidates:
        if _try_download(BRICKLINK_URL.format(color=cand, ref=ref), path):
            return path

    # 2) Variantes con sufijo a/b/c
    for suffix in ("a", "b", "c"):
        ref_v = f"{ref}{suffix}"
        for cand in bl_candidates[:5]:
            if _try_download(BRICKLINK_URL.format(color=cand, ref=ref_v), path):
                return path

    # 3) Rebrickable LDraw render
    ldraw_candidates: list = []
    if nm_low and nm_low in NAME_TO_LDRAW_ID:
        ldraw_candidates.append(NAME_TO_LDRAW_ID[nm_low])
    for ld in ("4", "0", "71", "72", "15", "14", "36", "47", "70", "1"):
        if ld not in ldraw_candidates:
            ldraw_candidates.append(ld)
    for ld in ldraw_candidates:
        if _try_download(REBRICKABLE_LDRAW_URL.format(color=ld, ref=ref), path):
            return path

    print(f"   [WARN] No se pudo descargar imagen para {ref} "
          f"(color={color_code}, name={color_name})")
    return None


def prepare_image_for_excel(src_path: str,
                            target_height: int = IMG_HEIGHT) -> XLImage:
    """Carga la imagen, la redimensiona y la prepara para openpyxl."""
    img = PILImage.open(src_path).convert("RGBA")
    bg = PILImage.new("RGBA", img.size, (255, 255, 255, 255))
    bg.paste(img, mask=img.split()[-1])
    img = bg.convert("RGB")
    w, h  = img.size
    scale = target_height / float(h)
    new_w, new_h = int(w * scale), int(h * scale)
    img = img.resize((new_w, new_h), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    xl_img = XLImage(buf)
    xl_img.width  = new_w
    xl_img.height = new_h
    return xl_img


# ══════════════════════════════════════════════════════════════════════════════
#  Construcción del Excel
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FILL_BLUE = PatternFill("solid", fgColor="1A237E")
HEADER_FONT      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CELL_FONT        = Font(name="Calibri", size=11)
CODE_FONT        = Font(name="Consolas", size=11, bold=True)
THIN             = Side(border_style="thin", color="CCCCCC")
BORDER           = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Colores de relleno para distinguir el tipo de pieza
FILL_REGULAR = PatternFill("solid", fgColor="FFFFFF")
FILL_MINIFIG = PatternFill("solid", fgColor="FFF8E1")  # amarillo muy suave


def build_excel():
    print(f"[Excel] Set {SET_ID} — leyendo BD...")
    inv = fetch_set_inventory(SET_ID)
    pose_counts = fetch_pose_counts(SET_ID)

    set_name = inv["set_name"]
    parts    = inv["regular_parts"] + inv["minifig_parts"]
    n_reg    = len(inv["regular_parts"])
    n_mfp    = len(inv["minifig_parts"])

    print(f"[Excel] Set name: {set_name}")
    print(f"[Excel] Inventario: {n_reg} regulares + {n_mfp} de minifig "
          f"= {len(parts)} filas")
    print(f"[Excel] Refs con poses estables en BD: {len(pose_counts)}")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Set {SET_ID}"

    headers = ["Imagen", "Código", "Tipo", "Color", "Nombre",
               "Poses estables (BD)", "Esperadas"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL_BLUE
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    col_widths = {1: 16, 2: 12, 3: 18, 4: 24, 5: 38, 6: 20, 7: 12}
    for idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[1].height = 26

    # 3. Filas
    for i, part in enumerate(parts, start=2):
        ref        = str(part["ref"])
        color_code = str(part.get("color_code", "0"))
        color_name = part.get("color_name", "") or ""
        name       = part.get("name", "") or ""
        n_poses    = pose_counts.get(ref, 0)
        source     = part.get("source", "regular")
        mref       = part.get("minifig_ref")

        if source == "minifig":
            tipo = f"Minifig ({mref})"
            row_fill = FILL_MINIFIG
        else:
            tipo = "Regular" + (f" + Minifig ({mref})" if mref else "")
            row_fill = FILL_REGULAR

        print(f"  [{i-1:02d}/{len(parts)}] {tipo:25s}  {ref:14s}  "
              f"color={color_code:>3} ({color_name})  poses={n_poses}")

        ws.row_dimensions[i].height = IMG_HEIGHT * 0.78

        # Imagen
        img_path = download_bricklink_image(ref, color_code, color_name)
        if img_path:
            try:
                xl_img = prepare_image_for_excel(img_path, IMG_HEIGHT)
                ws.add_image(xl_img, f"A{i}")
            except Exception as e:
                print(f"     [WARN] No se pudo incrustar imagen: {e}")

        # Imagen (col 1) — pintamos relleno + borde aunque la imagen flote
        cell_img = ws.cell(row=i, column=1)
        cell_img.fill   = row_fill
        cell_img.border = BORDER

        # Código
        c = ws.cell(row=i, column=2, value=ref)
        c.font = CODE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.fill = row_fill

        # Tipo
        c = ws.cell(row=i, column=3, value=tipo)
        c.font = Font(name="Calibri", size=10, bold=(source == "minifig"),
                      color=("8D6E00" if source == "minifig" else "1A237E"))
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        c.fill = row_fill

        # Color
        c = ws.cell(row=i, column=4,
                    value=f"{color_code} — {color_name}".strip(" —"))
        c.font = CELL_FONT
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        c.fill = row_fill

        # Nombre
        c = ws.cell(row=i, column=5, value=name)
        c.font = CELL_FONT
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border = BORDER
        c.fill = row_fill

        # Poses estables (con badge "—" si la pieza es de minifig porque no
        # se simulan poses para piezas-minifig — no aplica)
        if source == "minifig":
            display = "—"
            font = Font(name="Calibri", size=11, bold=True, color="888888",
                        italic=True)
        else:
            display = n_poses
            font = Font(name="Calibri", size=11, bold=True,
                        color=("2E7D32" if n_poses > 0 else "B71C1C"))
        c = ws.cell(row=i, column=6, value=display)
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.fill = row_fill

        # Esperadas (vacía)
        c = ws.cell(row=i, column=7, value=None)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        c.fill = row_fill

    ws.freeze_panes = "A2"

    # Hoja de leyenda
    ws2 = wb.create_sheet("Leyenda")
    ws2["A1"] = f"Set {SET_ID}: {set_name}"
    ws2["A1"].font = Font(bold=True, size=14, color="1A237E")
    ws2["A3"] = f"Inventario: {n_reg} piezas regulares + {n_mfp} piezas de minifig"
    ws2["A4"] = f"Total filas en hoja principal: {len(parts)}"
    ws2["A6"] = "Tipo \"Regular\""
    ws2["A6"].font = Font(bold=True, color="1A237E")
    ws2["B6"] = ("Pieza del inventario regular del set "
                 "(la que va sobre la cinta transportadora).")
    ws2["A7"] = "Tipo \"Minifig (sw...)\""
    ws2["A7"].font = Font(bold=True, color="8D6E00")
    ws2["B7"] = ("Pieza que pertenece al inventario interno de una minifigura "
                 "(cabeza, torso, brazo, pierna, accesorio…). "
                 "No se cuentan poses estables individuales.")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 80

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"\n[Excel] ✓ Guardado: {OUT_FILE}")
    print(f"[Excel]   {len(parts)} filas · cache imágenes: {CACHE_DIR}")


if __name__ == "__main__":
    build_excel()
